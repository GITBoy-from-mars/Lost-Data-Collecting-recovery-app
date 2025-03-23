
import os
import pandas as pd
import xml.etree.ElementTree as ET
from collections import defaultdict
import openpyxl

# ----------------------- Step 1: Convert XML to Excel -----------------------
def parse_node(node, path="", parent_data=None):
    if node is None:
        return []
    
    parent_data = parent_data.copy() if parent_data else {}
    results = []
    
    # Process attributes
    for attr, value in node.attrib.items():
        parent_data[f"{path}@{attr}"] = value
    
    # Process children
    has_children = False
    for child in node:
        has_children = True
        child_path = f"{path}/{child.tag}" if path else child.tag
        results.extend(parse_node(child, child_path, parent_data))
    
    # Process leaf node
    if not has_children and node.text and node.text.strip():
        parent_data[path] = node.text.strip()
        return [parent_data]
    
    return results

def odk_xml_to_excel(xml_file, excel_file):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        all_data = parse_node(root)
        
        df = pd.DataFrame(all_data)
        
        # Clean column names (remove path prefixes)
        df.columns = [col.split("/")[-1] for col in df.columns]
        
        df.to_excel(excel_file, index=False, engine='openpyxl')
        print(f"Successfully converted to {excel_file}")
        
    except Exception as e:
        print(f"Error: {str(e)}")

def convert_xml_folder(input_folder, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    for file_name in os.listdir(input_folder):
        if file_name.lower().endswith(".xml"):
            xml_file = os.path.join(input_folder, file_name)
            excel_file = os.path.join(output_folder, f"{os.path.splitext(file_name)[0]}.xlsx")
            print(f"Converting '{xml_file}' to '{excel_file}'...")
            odk_xml_to_excel(xml_file, excel_file)

# ----------------------- Step 2: Process Excel Sheets -----------------------
def shift_and_truncate_sheet(ws):
    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(3, max_col + 1):
        values = []
        for row in range(2, max_row + 1):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is not None:
                values.append(cell_val)
        
        for i, val in enumerate(values):
            ws.cell(row=2 + i, column=col).value = val
        
        for row in range(2 + len(values), max_row + 1):
            ws.cell(row=row, column=col).value = None

    max_data_row = 1
    for col in range(3, max_col + 1):
        for row in range(2, max_row + 1):
            if ws.cell(row=row, column=col).value is not None:
                max_data_row = max(max_data_row, row)

    for col in [1, 2]:
        for row in range(max_data_row + 1, max_row + 1):
            ws.cell(row=row, column=col).value = None

def process_workbook(input_file_path, output_file_path):
    wb = openpyxl.load_workbook(input_file_path)
    
    for ws in wb.worksheets:
        shift_and_truncate_sheet(ws)
    
    wb.save(output_file_path)

def process_folder(input_folder, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    for file_name in os.listdir(input_folder):
        if file_name.lower().endswith(".xlsx"):
            input_file_path = os.path.join(input_folder, file_name)
            output_file_path = os.path.join(output_folder, file_name)
            print(f"Processing: {input_file_path} -> {output_file_path}")
            process_workbook(input_file_path, output_file_path)
    
    print("All files processed successfully!")

# ----------------------- Step 3: Combine All Excel Sheets -----------------------
def combine_excels(folder_path, combined_file):
    combined_data = pd.DataFrame()
    header_included = False

    for file in os.listdir(folder_path):
        if file.endswith('.xlsx') or file.endswith('.xls'):
            file_path = os.path.join(folder_path, file)
            
            try:
                # Read the Excel file
                df = pd.read_excel(file_path)
                
                # If header is not included, add it
                if not header_included:
                    combined_data = pd.concat([combined_data, df], ignore_index=True)
                    header_included = True
                else:
                    # Append all data rows (header is already excluded by pd.read_excel)
                    combined_data = pd.concat([combined_data, df], ignore_index=True)

                print(f"Data from '{file}' added successfully.")
            
            except Exception as e:
                print(f"Failed to read '{file}': {e}")

    # Save combined data to a new Excel file
    combined_data.to_excel(combined_file, index=False)



# ----------------------- Step 4: Run All Steps -----------------------
if __name__ == "__main__":

    # Paths
    for i in range (6,14):

        # os.makedirs(f"S:\\Downloads\\processed_Outputs_of_path{i}", exist_ok=True)

        input_folder = f"S:\\Desktop\\path{i}"   # Folder containing XML files
        xml_to_excel_output = f"{input_folder}\\Outputs_of_path{i}"     # Output folder for XML to Excel
        
        processed_folder = f"{input_folder}\\processed_Outputs_of_path{i}" # Output folder for processed Excel files - removing gap
        combined_file = f"{xml_to_excel_output}\\combined result.xlsx" # Final combined  combined

        print("\n=== Step 1: Converting XML to Excel ===")
        convert_xml_folder(input_folder, xml_to_excel_output)

        print("\n=== Step 2: Processing Excel Files ===")
        process_folder(xml_to_excel_output, processed_folder)

        print("\n=== Step 3: Combining Excel Files ===")
        combine_excels(processed_folder, combined_file)

        print("\n=== All Steps Completed Successfully! ===")

